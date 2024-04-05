from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from nees_plots import commit_plots, lines_plots, issues_plots, avg_issue_plots


def file_generator(wb, sheet_name, pivot_df, df, metric, project_name):
    # Criando a sheet para o tipo de métrica especificado
    ws = wb.create_sheet(title=sheet_name)

    # Identificar o índice da coluna "nome" se presente
    nome_col_idx = None
    if 'Nome' in pivot_df.columns:
        nome_col_idx = pivot_df.columns.get_loc('Nome') + 1  # Ajustando para índice base-1 usado no Excel

    # Cálculo preliminar para o ajuste de largura das colunas baseado no cabeçalho
    col_widths = {}
    for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Formatação do cabeçalho
                cell.font = Font(bold=True, color="FFFFFF", size=20)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color='0072B2', end_color='0072B2', fill_type="solid")
            else:  # Alinhamento dos valores para a direita e fonte para os dados
                cell.font = Font(size=16)
                cell.alignment = Alignment(horizontal="left" if c_idx == nome_col_idx else "right")

            # Atualizando a largura máxima necessária para a coluna
            text_length = len(str(value))
            if c_idx in col_widths:
                col_widths[c_idx] = max(col_widths[c_idx], text_length)
            else:
                col_widths[c_idx] = text_length

    # Ajustando a largura das colunas com base no texto mais longo
    for col_idx, width in col_widths.items():
        # Vamos tentar aumentar o fator de ajuste para 1.5 e adicionar uma folga maior
        adjusted_width = (width * 1.5) + 3  # Aumentando a folga
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = adjusted_width

    function_map = {
        'lines_modified': (lines_plots, [df, project_name]),
        'commits': (commit_plots, [df, project_name]),
        'avg_issue': (avg_issue_plots, [df, project_name]),
        'issues_assigned': (issues_plots, [df, project_name]),
        }

    if metric in function_map:
        function_choice, args = function_map[metric]
        function_choice(*args)  # Chamar a função com argumentos desempacotados
    else:
        print("Opção inválida.")

    plot_png = {
        'lines_modified': '{}_linhas_modificadas_grafico.png'.format(project_name),
        'commits': '{}_commits_grafico.png'.format(project_name),
        'avg_issue': '{}_issues_avg_grafico.png'.format(project_name),
        'issues_assigned': '{}_issues_grafico.png'.format(project_name),
        }

    if metric in plot_png:
        img_name = plot_png[metric]
    else:
        print("Opção inválida.")

    # Inserindo a imagem abaixo da tabela
    img = Image(img_name)

    # Calcular a próxima linha vazia após a tabela para inserir a imagem
    next_row = ws.max_row + 2
    ws.add_image(img, f"A{next_row}")
